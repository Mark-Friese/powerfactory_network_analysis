"""
Excel report generator for PowerFactory analysis results.
"""

from typing import List, Dict, Any, Optional, Union
from pathlib import Path
from datetime import datetime
import pandas as pd

from ..core.results_manager import ResultsManager
from ..models.violation import Violation
from ..models.analysis_result import AnalysisType
from ..utils.logger import AnalysisLogger

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    from openpyxl.chart import BarChart, Reference
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False


class ExcelReporter:
    """
    Excel report generator for PowerFactory network analysis results.
    
    Creates comprehensive Excel reports with multiple sheets for different
    analysis aspects including violations, asset loading, and summaries.
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize Excel reporter.
        
        Args:
            config: Reporter configuration
        """
        self.logger = AnalysisLogger(self.__class__.__name__)
        self.config = config or {}
        
        if not EXCEL_AVAILABLE:
            self.logger.error("Excel dependencies not available. Install openpyxl and pandas.")
            raise ImportError("Excel reporting requires openpyxl and pandas")
        
        # Excel formatting styles
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.violation_fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
        self.warning_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        self.logger.info("Excel reporter initialized")
    
    def generate_report(self, results_manager: ResultsManager, 
                       output_path: Union[str, Path]) -> bool:
        """
        Generate complete Excel report.
        
        Args:
            results_manager: Results manager with analysis data
            output_path: Path for output Excel file
            
        Returns:
            True if successful
        """
        try:
            self.logger.info(f"Generating Excel report: {output_path}")
            
            # Create workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Generate sheets
            self._create_summary_sheet(workbook, results_manager)
            self._create_violations_sheet(workbook, results_manager)
            self._create_thermal_analysis_sheet(workbook, results_manager)
            self._create_voltage_analysis_sheet(workbook, results_manager)
            self._create_contingency_summary_sheet(workbook, results_manager)
            self._create_asset_loading_sheet(workbook, results_manager)
            
            # Add charts if configured
            if self.config.get('excel', {}).get('include_charts', True):
                self._add_charts(workbook, results_manager)
            
            # Save workbook
            workbook.save(output_path)
            self.logger.info(f"Excel report generated successfully: {output_path}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Failed to generate Excel report: {e}")
            return False
    
    def _create_summary_sheet(self, workbook: openpyxl.Workbook, 
                             results_manager: ResultsManager) -> None:
        """Create executive summary sheet."""
        sheet = workbook.create_sheet("Executive Summary")
        
        # Title
        sheet['A1'] = "PowerFactory Network Analysis Report"
        sheet['A1'].font = Font(size=16, bold=True)
        sheet['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary statistics
        stats = results_manager.get_summary_statistics()
        
        row = 4
        sheet[f'A{row}'] = "Analysis Summary"
        sheet[f'A{row}'].font = self.header_font
        sheet[f'A{row}'].fill = self.header_fill
        row += 1
        
        summary_data = [
            ["Total Violations", stats.get('total_violations', 0)],
            ["Base Case Violations", stats.get('base_case_violations', 0)],
            ["Contingency Violations", stats.get('contingency_violations', 0)],
            ["Thermal Violations", stats.get('thermal_violations', 0)],
            ["Voltage Violations", stats.get('voltage_violations', 0)]
        ]
        
        for item, value in summary_data:
            sheet[f'A{row}'] = item
            sheet[f'B{row}'] = value
            row += 1
        
        # Severity breakdown
        row += 2
        sheet[f'A{row}'] = "Violation Severity"
        sheet[f'A{row}'].font = self.header_font
        sheet[f'A{row}'].fill = self.header_fill
        row += 1
        
        severity = stats.get('severity_breakdown', {})
        for level, count in severity.items():
            sheet[f'A{row}'] = level.capitalize()
            sheet[f'B{row}'] = count
            if level == 'critical' and count > 0:
                sheet[f'B{row}'].fill = self.violation_fill
            row += 1
        
        # Regional breakdown
        row += 2
        sheet[f'A{row}'] = "Regional Breakdown"
        sheet[f'A{row}'].font = self.header_font
        sheet[f'A{row}'].fill = self.header_fill
        row += 1
        
        regional = stats.get('regional_breakdown', {})
        for region, count in regional.items():
            sheet[f'A{row}'] = region.capitalize()
            sheet[f'B{row}'] = count
            row += 1
        
        # Auto-fit columns
        self._autofit_columns(sheet)
    
    def _create_violations_sheet(self, workbook: openpyxl.Workbook, 
                                results_manager: ResultsManager) -> None:
        """Create violations summary sheet."""
        sheet = workbook.create_sheet("Violations")
        
        violations = results_manager.get_all_violations()
        
        if not violations:
            sheet['A1'] = "No violations found"
            return
        
        # Create DataFrame
        violation_data = []
        for violation in violations:
            violation_data.append({
                'Element Name': violation.element_name,
                'Element Type': violation.element_type.value,
                'Voltage Level (kV)': violation.voltage_level,
                'Region': violation.region.value,
                'Analysis Type': violation.analysis_type.value,
                'Violation Value': violation.violation_value,
                'Limit Value': violation.limit_value,
                'Severity': violation.severity,
                'Scenario': violation.scenario,
                'Violation Type': violation.metadata.get('violation_type', 'N/A')
            })
        
        df = pd.DataFrame(violation_data)
        
        # Write to sheet
        self._write_dataframe_to_sheet(sheet, df, "Network Violations")
        
        # Apply conditional formatting
        self._apply_violation_formatting(sheet, len(df) + 2)
    
    def _create_thermal_analysis_sheet(self, workbook: openpyxl.Workbook, 
                                      results_manager: ResultsManager) -> None:
        """Create thermal analysis results sheet."""
        sheet = workbook.create_sheet("Thermal Analysis")
        
        # Get thermal violations
        thermal_violations = results_manager.get_violations_by_type(AnalysisType.THERMAL)
        
        if not thermal_violations:
            sheet['A1'] = "No thermal violations found"
            return
        
        # Create DataFrame
        thermal_data = []
        for violation in thermal_violations:
            thermal_data.append({
                'Element Name': violation.element_name,
                'Element Type': violation.element_type.value,
                'Voltage Level (kV)': violation.voltage_level,
                'Region': violation.region.value,
                'Loading (%)': violation.violation_value,
                'Limit (%)': violation.limit_value,
                'Overload (%)': violation.violation_value - violation.limit_value,
                'Severity': violation.severity,
                'Scenario': violation.scenario,
                'Current (A)': violation.metadata.get('current_amps', 'N/A'),
                'Power (MW)': violation.metadata.get('power_mw', 'N/A')
            })
        
        df = pd.DataFrame(thermal_data)
        
        # Sort by overload percentage (descending)
        df = df.sort_values('Overload (%)', ascending=False)
        
        # Write to sheet
        self._write_dataframe_to_sheet(sheet, df, "Thermal Loading Violations")
        
        # Apply formatting
        self._apply_thermal_formatting(sheet, len(df) + 2)
    
    def _create_voltage_analysis_sheet(self, workbook: openpyxl.Workbook, 
                                      results_manager: ResultsManager) -> None:
        """Create voltage analysis results sheet."""
        sheet = workbook.create_sheet("Voltage Analysis")
        
        # Get voltage violations
        voltage_violations = results_manager.get_violations_by_type(AnalysisType.VOLTAGE)
        
        if not voltage_violations:
            sheet['A1'] = "No voltage violations found"
            return
        
        # Create DataFrame
        voltage_data = []
        for violation in voltage_violations:
            voltage_data.append({
                'Bus Name': violation.element_name,
                'Voltage Level (kV)': violation.voltage_level,
                'Region': violation.region.value,
                'Voltage (pu)': violation.violation_value,
                'Limit (pu)': violation.limit_value,
                'Deviation (pu)': abs(violation.violation_value - violation.limit_value),
                'Violation Type': violation.metadata.get('violation_type', 'N/A'),
                'Severity': violation.severity,
                'Scenario': violation.scenario,
                'Voltage (kV)': violation.metadata.get('voltage_kv', 'N/A'),
                'Angle (deg)': violation.metadata.get('angle_deg', 'N/A')
            })
        
        df = pd.DataFrame(voltage_data)
        
        # Sort by deviation (descending)
        df = df.sort_values('Deviation (pu)', ascending=False)
        
        # Write to sheet
        self._write_dataframe_to_sheet(sheet, df, "Voltage Violations")
        
        # Apply formatting
        self._apply_voltage_formatting(sheet, len(df) + 2)
    
    def _create_contingency_summary_sheet(self, workbook: openpyxl.Workbook, 
                                         results_manager: ResultsManager) -> None:
        """Create contingency analysis summary sheet."""
        sheet = workbook.create_sheet("Contingency Summary")
        
        worst_contingencies = results_manager.get_worst_contingencies(20)
        
        if not worst_contingencies:
            sheet['A1'] = "No contingency violations found"
            return
        
        # Create DataFrame
        df = pd.DataFrame(worst_contingencies)
        
        # Write to sheet
        self._write_dataframe_to_sheet(sheet, df, "Worst Contingencies (Top 20)")
        
        # Apply formatting
        self._apply_contingency_formatting(sheet, len(df) + 2)
    
    def _create_asset_loading_sheet(self, workbook: openpyxl.Workbook, 
                                   results_manager: ResultsManager) -> None:
        """Create asset loading summary sheet."""
        sheet = workbook.create_sheet("Asset Loading")
        
        asset_summary = results_manager.get_asset_loading_summary()
        
        if not asset_summary:
            sheet['A1'] = "No asset loading data available"
            return
        
        # Loading summary
        row = 1
        sheet[f'A{row}'] = "Asset Loading Summary"
        sheet[f'A{row}'].font = self.header_font
        sheet[f'A{row}'].fill = self.header_fill
        row += 2
        
        summary_items = [
            ("Total Elements Analyzed", asset_summary.get('total_elements', 0)),
            ("Maximum Loading (%)", f"{asset_summary.get('max_loading', 0):.1f}"),
            ("Average Loading (%)", f"{asset_summary.get('avg_loading', 0):.1f}"),
            ("Elements >90% Loading", asset_summary.get('elements_over_90', 0)),
            ("Elements >100% Loading", asset_summary.get('elements_over_100', 0))
        ]
        
        for item, value in summary_items:
            sheet[f'A{row}'] = item
            sheet[f'B{row}'] = value
            row += 1
        
        # Loading distribution
        row += 2
        sheet[f'A{row}'] = "Loading Distribution"
        sheet[f'A{row}'].font = self.header_font
        sheet[f'A{row}'].fill = self.header_fill
        row += 1
        
        distribution = asset_summary.get('loading_distribution', {})
        for range_str, count in distribution.items():
            sheet[f'A{row}'] = range_str
            sheet[f'B{row}'] = count
            if range_str == '>100%' and count > 0:
                sheet[f'B{row}'].fill = self.violation_fill
            elif range_str == '90-100%' and count > 0:
                sheet[f'B{row}'].fill = self.warning_fill
            row += 1
        
        self._autofit_columns(sheet)
    
    def _write_dataframe_to_sheet(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                 df: pd.DataFrame, title: str) -> None:
        """Write DataFrame to Excel sheet with formatting."""
        # Title
        sheet['A1'] = title
        sheet['A1'].font = Font(size=14, bold=True)
        
        # Headers
        start_row = 3
        for col_num, column_title in enumerate(df.columns, 1):
            cell = sheet.cell(row=start_row, column=col_num)
            cell.value = column_title
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
        
        # Data
        for row_num, row_data in enumerate(df.values, start_row + 1):
            for col_num, value in enumerate(row_data, 1):
                cell = sheet.cell(row=row_num, column=col_num)
                cell.value = value
                cell.border = self.border
                
                # Format numbers
                if isinstance(value, (int, float)) and col_num > 2:
                    if 'percentage' in df.columns[col_num-1].lower() or '%' in str(df.columns[col_num-1]):
                        cell.number_format = '0.1'
                    else:
                        cell.number_format = '0.00'
        
        self._autofit_columns(sheet)
    
    def _apply_violation_formatting(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                   data_rows: int) -> None:
        """Apply conditional formatting for violations."""
        severity_col = None
        
        # Find severity column
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=3, column=col).value == 'Severity':
                severity_col = col
                break
        
        if severity_col:
            for row in range(4, data_rows + 2):
                severity = sheet.cell(row=row, column=severity_col).value
                if severity == 'Critical':
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = self.violation_fill
                elif severity == 'High':
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = self.warning_fill
    
    def _apply_thermal_formatting(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                 data_rows: int) -> None:
        """Apply formatting specific to thermal analysis."""
        self._apply_violation_formatting(sheet, data_rows)
        
        # Highlight overloaded elements
        overload_col = None
        for col in range(1, sheet.max_column + 1):
            if 'Overload' in str(sheet.cell(row=3, column=col).value):
                overload_col = col
                break
        
        if overload_col:
            for row in range(4, data_rows + 2):
                overload = sheet.cell(row=row, column=overload_col).value
                if isinstance(overload, (int, float)) and overload > 20:
                    sheet.cell(row=row, column=overload_col).fill = self.violation_fill
                elif isinstance(overload, (int, float)) and overload > 10:
                    sheet.cell(row=row, column=overload_col).fill = self.warning_fill
    
    def _apply_voltage_formatting(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                 data_rows: int) -> None:
        """Apply formatting specific to voltage analysis."""
        self._apply_violation_formatting(sheet, data_rows)
    
    def _apply_contingency_formatting(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                                     data_rows: int) -> None:
        """Apply formatting specific to contingency summary."""
        # Highlight contingencies with critical violations
        critical_col = None
        for col in range(1, sheet.max_column + 1):
            if 'critical_violations' in str(sheet.cell(row=3, column=col).value):
                critical_col = col
                break
        
        if critical_col:
            for row in range(4, data_rows + 2):
                critical_count = sheet.cell(row=row, column=critical_col).value
                if isinstance(critical_count, (int, float)) and critical_count > 0:
                    for col in range(1, sheet.max_column + 1):
                        sheet.cell(row=row, column=col).fill = self.violation_fill
    
    def _add_charts(self, workbook: openpyxl.Workbook, results_manager: ResultsManager) -> None:
        """Add charts to the workbook."""
        try:
            # Add chart to summary sheet if it exists
            if "Executive Summary" in workbook.sheetnames:
                summary_sheet = workbook["Executive Summary"]
                self._add_severity_chart(summary_sheet, results_manager)
                
        except Exception as e:
            self.logger.warning(f"Could not add charts: {e}")
    
    def _add_severity_chart(self, sheet: openpyxl.worksheet.worksheet.Worksheet, 
                           results_manager: ResultsManager) -> None:
        """Add severity breakdown chart."""
        try:
            stats = results_manager.get_summary_statistics()
            severity = stats.get('severity_breakdown', {})
            
            if not severity:
                return
            
            # Find a good location for the chart
            chart_start_row = 20
            chart_start_col = 4
            
            # Add data for chart
            data_start_row = chart_start_row
            sheet[f'{chr(65 + chart_start_col)}{data_start_row}'] = "Severity"
            sheet[f'{chr(66 + chart_start_col)}{data_start_row}'] = "Count"
            
            for i, (severity_level, count) in enumerate(severity.items(), 1):
                sheet[f'{chr(65 + chart_start_col)}{data_start_row + i}'] = severity_level.capitalize()
                sheet[f'{chr(66 + chart_start_col)}{data_start_row + i}'] = count
            
            # Create chart
            chart = BarChart()
            chart.title = "Violation Severity Breakdown"
            chart.y_axis.title = "Number of Violations"
            chart.x_axis.title = "Severity Level"
            
            data = Reference(sheet, 
                           min_col=chart_start_col + 2, 
                           min_row=data_start_row,
                           max_col=chart_start_col + 2, 
                           max_row=data_start_row + len(severity))
            categories = Reference(sheet, 
                                 min_col=chart_start_col + 1, 
                                 min_row=data_start_row + 1,
                                 max_col=chart_start_col + 1, 
                                 max_row=data_start_row + len(severity))
            
            chart.add_data(data, titles_from_data=True)
            chart.set_categories(categories)
            
            # Add chart to sheet
            sheet.add_chart(chart, f'{chr(65 + chart_start_col + 4)}{chart_start_row}')
            
        except Exception as e:
            self.logger.debug(f"Could not add severity chart: {e}")
    
    def _autofit_columns(self, sheet: openpyxl.worksheet.worksheet.Worksheet) -> None:
        """Auto-fit column widths."""
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            sheet.column_dimensions[column_letter].width = adjusted_width
