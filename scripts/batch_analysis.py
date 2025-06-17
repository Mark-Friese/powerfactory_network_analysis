#!/usr/bin/env python3
"""
Batch analysis script for PowerFactory network analysis.

This script enables batch processing of multiple PowerFactory studies,
configurations, or scenarios for comparative analysis.

Usage:
    python batch_analysis.py [options]

Examples:
    # Run batch analysis with study list
    python batch_analysis.py --studies studies.yaml

    # Run with multiple configurations
    python batch_analysis.py --configs config1.yaml config2.yaml

    # Process all .pfd files in directory
    python batch_analysis.py --study-dir ./studies --pattern "*.pfd"
"""

import argparse
import sys
import traceback
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Any, Optional, Tuple
import time
import json

# Add src to Python path
current_dir = Path(__file__).parent
src_dir = current_dir.parent / "src"
sys.path.insert(0, str(src_dir))

from src.core.network_analyzer import NetworkAnalyzer
from src.core.results_manager import ResultsManager
from src.reports.excel_reporter import ExcelReporter
from src.reports.csv_reporter import CSVReporter
from src.utils.logger import AnalysisLogger, get_logger
from src.utils.file_handler import FileHandler
from src.utils.validation import InputValidator


class BatchAnalysisManager:
    """
    Manager for batch analysis operations.
    
    Handles multiple study processing, result aggregation,
    and comparative reporting.
    """
    
    def __init__(self, config: Dict[str, Any]):
        """Initialize batch analysis manager."""
        self.config = config
        self.logger = AnalysisLogger(self.__class__.__name__)
        self.file_handler = FileHandler()
        self.validator = InputValidator()
        
        # Batch results storage
        self.batch_results: Dict[str, Dict[str, Any]] = {}
        self.failed_studies: List[Dict[str, str]] = []
        
        self.logger.info("Batch analysis manager initialized")
    
    def load_study_list(self, study_list_path: str) -> Optional[List[Dict[str, Any]]]:
        """Load study list from YAML file."""
        try:
            study_data = self.file_handler.read_yaml(study_list_path)
            if not study_data or 'studies' not in study_data:
                self.logger.error("Invalid study list format")
                return None
            
            studies = study_data['studies']
            self.logger.info(f"Loaded {len(studies)} studies from {study_list_path}")
            return studies
            
        except Exception as e:
            self.logger.error(f"Failed to load study list: {e}")
            return None
    
    def discover_studies(self, study_dir: str, pattern: str = "*.pfd") -> List[Dict[str, Any]]:
        """Discover studies in directory."""
        try:
            study_path = Path(study_dir)
            if not study_path.exists():
                self.logger.error(f"Study directory does not exist: {study_dir}")
                return []
            
            study_files = list(study_path.glob(pattern))
            studies = []
            
            for study_file in study_files:
                study = {
                    'name': study_file.stem,
                    'path': str(study_file),
                    'description': f"Auto-discovered study: {study_file.name}"
                }
                studies.append(study)
            
            self.logger.info(f"Discovered {len(studies)} studies in {study_dir}")
            return studies
            
        except Exception as e:
            self.logger.error(f"Failed to discover studies: {e}")
            return []
    
    def create_study_configs(self, base_config: Dict[str, Any], 
                           config_variations: List[str]) -> List[Tuple[str, Dict[str, Any]]]:
        """Create study configurations from base config and variations."""
        try:
            study_configs = []
            
            for config_path in config_variations:
                # Load variation config
                variation_config = self.file_handler.read_yaml(config_path)
                if variation_config is None:
                    self.logger.warning(f"Could not load config variation: {config_path}")
                    continue
                
                # Merge with base config
                merged_config = self._merge_configs(base_config, variation_config)
                
                # Create study name from config filename
                study_name = Path(config_path).stem
                study_configs.append((study_name, merged_config))
            
            self.logger.info(f"Created {len(study_configs)} study configurations")
            return study_configs
            
        except Exception as e:
            self.logger.error(f"Failed to create study configs: {e}")
            return []
    
    def _merge_configs(self, base_config: Dict[str, Any], 
                      override_config: Dict[str, Any]) -> Dict[str, Any]:
        """Merge configuration dictionaries."""
        merged = base_config.copy()
        
        def deep_merge(base_dict, override_dict):
            for key, value in override_dict.items():
                if (key in base_dict and 
                    isinstance(base_dict[key], dict) and 
                    isinstance(value, dict)):
                    deep_merge(base_dict[key], value)
                else:
                    base_dict[key] = value
        
        deep_merge(merged, override_config)
        return merged
    
    def run_batch_analysis(self, studies: List[Dict[str, Any]], 
                          output_dir: Path,
                          base_config: Optional[Dict[str, Any]] = None) -> bool:
        """Run batch analysis on list of studies."""
        try:
            total_studies = len(studies)
            successful_studies = 0
            
            self.logger.info(f"Starting batch analysis of {total_studies} studies")
            start_time = time.time()
            
            for i, study in enumerate(studies, 1):
                study_name = study.get('name', f'study_{i}')
                self.logger.info(f"Processing study {i}/{total_studies}: {study_name}")
                
                try:
                    # Run individual study analysis
                    study_result = self._analyze_single_study(study, base_config, output_dir)
                    
                    if study_result:
                        self.batch_results[study_name] = study_result
                        successful_studies += 1
                        self.logger.info(f"Study {study_name} completed successfully")
                    else:
                        self._record_failed_study(study_name, "Analysis failed")
                        
                except Exception as e:
                    self.logger.error(f"Study {study_name} failed: {e}")
                    self._record_failed_study(study_name, str(e))
            
            elapsed_time = time.time() - start_time
            
            self.logger.info(f"Batch analysis completed: {successful_studies}/{total_studies} successful")
            self.logger.info(f"Total time: {elapsed_time:.1f} seconds")
            
            return successful_studies > 0
            
        except Exception as e:
            self.logger.error(f"Batch analysis failed: {e}")
            return False
    
    def _analyze_single_study(self, study: Dict[str, Any], 
                             base_config: Optional[Dict[str, Any]],
                             output_dir: Path) -> Optional[Dict[str, Any]]:
        """Analyze a single study."""
        try:
            study_name = study.get('name', 'unnamed_study')
            
            # Use study-specific config if provided, otherwise use base config
            if 'config' in study:
                if isinstance(study['config'], str):
                    # Config file path
                    study_config = self.file_handler.read_yaml(study['config'])
                else:
                    # Inline config
                    study_config = study['config']
                
                if base_config:
                    study_config = self._merge_configs(base_config, study_config)
            else:
                study_config = base_config or {}
            
            # Initialize analyzer with study config
            analyzer = NetworkAnalyzer()
            analyzer.config = study_config
            
            # Set PowerFactory study if specified
            if 'path' in study and study['path'].endswith('.pfd'):
                # Note: Actual PowerFactory study switching would be implemented here
                # This is a placeholder for the PowerFactory-specific code
                self.logger.debug(f"Would switch to PowerFactory study: {study['path']}")
            
            # Run analysis
            results = analyzer.run_full_analysis()
            
            if results:
                # Add study metadata
                results['study_info'] = {
                    'name': study_name,
                    'description': study.get('description', ''),
                    'path': study.get('path', ''),
                    'config': study_config
                }
                
                # Save individual study results
                study_output_dir = output_dir / "studies" / study_name
                study_output_dir.mkdir(parents=True, exist_ok=True)
                
                results_file = study_output_dir / "results.json"
                self.file_handler.write_json(results, results_file)
                
                return results
            
            return None
            
        except Exception as e:
            self.logger.error(f"Single study analysis failed: {e}")
            return None
    
    def _record_failed_study(self, study_name: str, error_message: str) -> None:
        """Record failed study information."""
        self.failed_studies.append({
            'name': study_name,
            'error': error_message,
            'timestamp': datetime.now().isoformat()
        })
    
    def generate_comparative_report(self, output_dir: Path) -> bool:
        """Generate comparative report across all studies."""
        try:
            if not self.batch_results:
                self.logger.warning("No batch results available for comparative report")
                return False
            
            self.logger.info("Generating comparative report")
            
            # Create comparative analysis
            comparative_data = self._create_comparative_analysis()
            
            # Save comparative data
            comp_file = output_dir / "comparative_analysis.json"
            self.file_handler.write_json(comparative_data, comp_file)
            
            # Generate Excel comparative report
            try:
                self._generate_comparative_excel(comparative_data, output_dir)
            except Exception as e:
                self.logger.warning(f"Could not generate comparative Excel report: {e}")
            
            # Generate CSV comparative report
            try:
                self._generate_comparative_csv(comparative_data, output_dir)
            except Exception as e:
                self.logger.warning(f"Could not generate comparative CSV report: {e}")
            
            return True
            
        except Exception as e:
            self.logger.error(f"Comparative report generation failed: {e}")
            return False
    
    def _create_comparative_analysis(self) -> Dict[str, Any]:
        """Create comparative analysis across studies."""
        comparative = {
            'timestamp': datetime.now().isoformat(),
            'studies_analyzed': list(self.batch_results.keys()),
            'failed_studies': self.failed_studies,
            'summary': {},
            'violations_comparison': {},
            'performance_comparison': {}
        }
        
        # Aggregate violations across studies
        all_violations = []
        study_summaries = {}
        
        for study_name, results in self.batch_results.items():
            try:
                # Create results manager for this study
                results_manager = ResultsManager()
                results_manager.add_analysis_results(results)
                
                # Get study summary
                study_summary = results_manager.get_summary_statistics()
                study_summaries[study_name] = study_summary
                
                # Get violations
                violations = results_manager.get_all_violations()
                for violation in violations:
                    violation_dict = violation.to_dict()
                    violation_dict['study_name'] = study_name
                    all_violations.append(violation_dict)
                    
            except Exception as e:
                self.logger.warning(f"Could not process study {study_name} for comparison: {e}")
        
        # Create comparative summaries
        comparative['summary'] = study_summaries
        comparative['all_violations'] = all_violations
        
        # Performance comparison
        comparative['performance_comparison'] = self._compare_study_performance(study_summaries)
        
        return comparative
    
    def _compare_study_performance(self, study_summaries: Dict[str, Dict]) -> Dict[str, Any]:
        """Compare performance across studies."""
        comparison = {
            'best_performing_study': None,
            'worst_performing_study': None,
            'violation_trends': {},
            'rankings': {}
        }
        
        try:
            # Rank studies by total violations
            violation_counts = {}
            for study_name, summary in study_summaries.items():
                violation_counts[study_name] = summary.get('total_violations', 0)
            
            # Sort by violations (ascending - fewer violations is better)
            sorted_studies = sorted(violation_counts.items(), key=lambda x: x[1])
            
            if sorted_studies:
                comparison['best_performing_study'] = sorted_studies[0][0]
                comparison['worst_performing_study'] = sorted_studies[-1][0]
            
            # Create rankings
            comparison['rankings'] = {
                'by_total_violations': sorted_studies,
                'by_critical_violations': self._rank_by_metric(study_summaries, 'severity_breakdown.critical'),
                'by_thermal_violations': self._rank_by_metric(study_summaries, 'thermal_violations'),
                'by_voltage_violations': self._rank_by_metric(study_summaries, 'voltage_violations')
            }
            
        except Exception as e:
            self.logger.warning(f"Performance comparison failed: {e}")
        
        return comparison
    
    def _rank_by_metric(self, study_summaries: Dict[str, Dict], metric_path: str) -> List[Tuple[str, int]]:
        """Rank studies by specific metric."""
        try:
            metric_values = {}
            
            for study_name, summary in study_summaries.items():
                # Navigate nested dictionary path
                value = summary
                for key in metric_path.split('.'):
                    value = value.get(key, {})
                    if not isinstance(value, dict):
                        break
                
                if isinstance(value, (int, float)):
                    metric_values[study_name] = value
                else:
                    metric_values[study_name] = 0
            
            return sorted(metric_values.items(), key=lambda x: x[1])
            
        except Exception as e:
            self.logger.warning(f"Ranking by {metric_path} failed: {e}")
            return []
    
    def _generate_comparative_excel(self, comparative_data: Dict[str, Any], 
                                   output_dir: Path) -> None:
        """Generate comparative Excel report."""
        try:
            import pandas as pd
            import openpyxl
            from openpyxl.styles import Font, PatternFill
            
            # Create workbook
            workbook = openpyxl.Workbook()
            workbook.remove(workbook.active)
            
            # Summary sheet
            summary_sheet = workbook.create_sheet("Study Comparison")
            
            # Study summaries
            if comparative_data.get('summary'):
                summaries = comparative_data['summary']
                
                # Create summary table
                studies = list(summaries.keys())
                metrics = ['total_violations', 'thermal_violations', 'voltage_violations']
                
                # Headers
                summary_sheet['A1'] = "Study Name"
                for i, metric in enumerate(metrics, 2):
                    summary_sheet.cell(row=1, column=i).value = metric.replace('_', ' ').title()
                
                # Data
                for row, study in enumerate(studies, 2):
                    summary_sheet.cell(row=row, column=1).value = study
                    for col, metric in enumerate(metrics, 2):
                        value = summaries[study].get(metric, 0)
                        summary_sheet.cell(row=row, column=col).value = value
            
            # Rankings sheet
            if comparative_data.get('performance_comparison', {}).get('rankings'):
                rankings_sheet = workbook.create_sheet("Rankings")
                rankings = comparative_data['performance_comparison']['rankings']
                
                col = 1
                for ranking_type, ranking_data in rankings.items():
                    rankings_sheet.cell(row=1, column=col).value = ranking_type.replace('_', ' ').title()
                    
                    for row, (study, value) in enumerate(ranking_data, 2):
                        rankings_sheet.cell(row=row, column=col).value = f"{study} ({value})"
                    
                    col += 2  # Leave a gap between columns
            
            # Save workbook
            excel_path = output_dir / "comparative_report.xlsx"
            workbook.save(excel_path)
            self.logger.info(f"Comparative Excel report saved: {excel_path}")
            
        except ImportError:
            self.logger.warning("Excel reporting not available")
        except Exception as e:
            self.logger.error(f"Comparative Excel generation failed: {e}")
    
    def _generate_comparative_csv(self, comparative_data: Dict[str, Any], 
                                 output_dir: Path) -> None:
        """Generate comparative CSV reports."""
        try:
            csv_dir = output_dir / "csv_comparative"
            csv_dir.mkdir(exist_ok=True)
            
            # Study summaries CSV
            if comparative_data.get('summary'):
                summaries = comparative_data['summary']
                summary_data = []
                
                for study_name, summary in summaries.items():
                    summary_data.append({
                        'study_name': study_name,
                        'total_violations': summary.get('total_violations', 0),
                        'thermal_violations': summary.get('thermal_violations', 0),
                        'voltage_violations': summary.get('voltage_violations', 0),
                        'base_case_violations': summary.get('base_case_violations', 0),
                        'contingency_violations': summary.get('contingency_violations', 0)
                    })
                
                self.file_handler.write_csv(summary_data, csv_dir / "study_summaries.csv")
            
            # All violations CSV
            if comparative_data.get('all_violations'):
                violations_data = comparative_data['all_violations']
                self.file_handler.write_csv(violations_data, csv_dir / "all_violations.csv")
            
            self.logger.info(f"Comparative CSV reports saved: {csv_dir}")
            
        except Exception as e:
            self.logger.error(f"Comparative CSV generation failed: {e}")


def parse_arguments() -> argparse.Namespace:
    """Parse command line arguments."""
    parser = argparse.ArgumentParser(
        description="PowerFactory Batch Network Analysis Tool",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    
    # Study specification options (mutually exclusive)
    study_group = parser.add_mutually_exclusive_group(required=True)
    study_group.add_argument(
        "--studies",
        type=str,
        help="YAML file containing list of studies to analyze"
    )
    
    study_group.add_argument(
        "--study-dir",
        type=str,
        help="Directory containing PowerFactory study files"
    )
    
    study_group.add_argument(
        "--configs",
        nargs="+",
        help="List of configuration files to process"
    )
    
    # Additional options for study discovery
    parser.add_argument(
        "--pattern",
        type=str,
        default="*.pfd",
        help="File pattern for study discovery (default: *.pfd)"
    )
    
    # Configuration options
    parser.add_argument(
        "--base-config",
        type=str,
        help="Base configuration file for all studies"
    )
    
    # Output options
    parser.add_argument(
        "--output-dir", "-o",
        type=str,
        default="./batch_output",
        help="Output directory for batch results (default: ./batch_output)"
    )
    
    parser.add_argument(
        "--no-individual-reports",
        action="store_true",
        help="Skip individual study reports"
    )
    
    parser.add_argument(
        "--no-comparative-report",
        action="store_true",
        help="Skip comparative report generation"
    )
    
    # Analysis options
    parser.add_argument(
        "--parallel",
        action="store_true",
        help="Run studies in parallel (experimental)"
    )
    
    parser.add_argument(
        "--max-workers",
        type=int,
        default=2,
        help="Maximum parallel workers (default: 2)"
    )
    
    # Logging options
    parser.add_argument(
        "--log-level",
        choices=["DEBUG", "INFO", "WARNING", "ERROR"],
        default="INFO",
        help="Logging level (default: INFO)"
    )
    
    parser.add_argument(
        "--verbose", "-v",
        action="store_true",
        help="Enable verbose output"
    )
    
    return parser.parse_args()


def main() -> int:
    """Main entry point."""
    args = parse_arguments()
    
    # Setup logging
    log_level = "DEBUG" if args.verbose else args.log_level
    logger = get_logger("batch_analysis")
    
    try:
        logger.info("PowerFactory Batch Network Analysis Tool")
        
        # Load base configuration
        base_config = None
        if args.base_config:
            file_handler = FileHandler()
            base_config = file_handler.read_yaml(args.base_config)
            if base_config is None:
                logger.error("Failed to load base configuration")
                return 1
        
        # Setup output directory
        output_path = Path(args.output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Initialize batch manager
        batch_manager = BatchAnalysisManager(base_config or {})
        
        # Load studies based on input method
        studies = []
        
        if args.studies:
            # Load from study list file
            studies = batch_manager.load_study_list(args.studies)
            if not studies:
                return 1
                
        elif args.study_dir:
            # Discover studies in directory
            studies = batch_manager.discover_studies(args.study_dir, args.pattern)
            if not studies:
                logger.error("No studies found in directory")
                return 1
                
        elif args.configs:
            # Create studies from config variations
            study_configs = batch_manager.create_study_configs(base_config or {}, args.configs)
            studies = [{'name': name, 'config': config} for name, config in study_configs]
            if not studies:
                return 1
        
        # Run batch analysis
        if not batch_manager.run_batch_analysis(studies, output_path, base_config):
            logger.error("Batch analysis failed")
            return 1
        
        # Generate comparative report
        if not args.no_comparative_report:
            if not batch_manager.generate_comparative_report(output_path):
                logger.warning("Comparative report generation failed")
        
        logger.info("Batch analysis completed successfully")
        return 0
        
    except KeyboardInterrupt:
        logger.info("Batch analysis interrupted by user")
        return 130
    except Exception as e:
        logger.error(f"Batch analysis failed: {e}")
        if args.verbose:
            logger.error(f"Traceback: {traceback.format_exc()}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
